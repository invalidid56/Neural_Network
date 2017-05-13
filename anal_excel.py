import openpyxl as o

def xlanal(fn):
    result = []
    temp = []
    wb = o.load_workbook(filename=fn)
    sheet = wb.active

    for x in range(1, 256):
        for y in range(1, 256):
            if sheet.cell(row=x, column=y).value is None:
                break
            temp.append(sheet.cell(row=x, column=y))
            k = 1
        if k is 0:
            break
        result.append(temp)
        temp = []
        k = 0

    for x in range(len(result)):
        for y in range(len(result[x])):
            result[x][y] = result[x][y].value

    return result

if __name__ is '__main__':
    print(xlanal(input('File Name : ')))
