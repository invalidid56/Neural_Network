import pyreg
from openpyxl import Workbook as wb
from openpyxl import load_workbook

f = input('File Name :  ')
wbread = load_workbook(filename=f)
mysheet = wbread['DataSet']

k = 0
p_set = []
while(1):
    x = mysheet.cell(row=k, column=0)
    y = mysheet.cell(row=k, column=1)
    if x is None:
        break
    p_set.append([x, y])


c = mysheet.cell(row = 2, column=3)

#  무작위로 선정한 데이터세트를 이용한 선형회귀분석 예시

s = pyreg.SimpleLinearRegression()

